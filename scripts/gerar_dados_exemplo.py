"""
Gera um arquivo XLSX de exemplo no formato Empresa A para demonstração.
Uso: python scripts/gerar_dados_exemplo.py
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

DESTINO = ROOT / "data" / "edificio_exemplo" / "2026-05"
DESTINO.mkdir(parents=True, exist_ok=True)


def escrever_cabecalho(ws, cols, cor="2563EB"):
    fill = PatternFill("solid", fgColor=cor)
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")


def criar_xlsx():
    wb = openpyxl.Workbook()

    # ── Aba Receitas ──
    ws = wb.active
    ws.title = "Receitas"
    escrever_cabecalho(ws, ["Descrição", "Previsto (R$)", "Realizado (R$)"])
    receitas = [
        ("Taxa de Condomínio",    52800.00, 49200.00),
        ("Taxa Extra — Reforma",   8000.00,  7500.00),
        ("Aluguel Salão de Festas",  600.00,    400.00),
        ("Multas e Juros",              0.00,    320.00),
    ]
    for r, dados in enumerate(receitas, 2):
        ws.cell(r, 1, dados[0])
        ws.cell(r, 2, dados[1])
        ws.cell(r, 3, dados[2])

    # ── Aba Despesas ──
    ws2 = wb.create_sheet("Despesas")
    escrever_cabecalho(ws2, ["Categoria", "Valor (R$)"])
    despesas = [
        ("Limpeza e Conservação",  8200.00),
        ("Manutenção Predial",     6400.00),
        ("Energia Elétrica",       4800.00),
        ("Segurança / Portaria",  12000.00),
        ("Administração",          3500.00),
        ("Jardinagem",             1200.00),
        ("Seguro do Edifício",     2100.00),
        ("Elevadores",             1800.00),
        ("Água e Esgoto",          3200.00),
        ("Outros",                   750.00),
    ]
    for r, (cat, val) in enumerate(despesas, 2):
        ws2.cell(r, 1, cat)
        ws2.cell(r, 2, val)

    # ── Aba Resumo ──
    ws3 = wb.create_sheet("Resumo")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 16
    dados_resumo = [
        ("Campo",              "Valor"),
        ("Saldo Anterior",     15420.50),
        ("",                   ""),
        ("Inadimplência Valor", 4680.00),
        ("Inadimplência %",    9.5),
        ("Unidades Inadimpl.", 5),
    ]
    for r, (label, val) in enumerate(dados_resumo, 1):
        ws3.cell(r, 1, label)
        ws3.cell(r, 2, val)
    ws3.cell(1, 1).font = Font(bold=True)
    ws3.cell(1, 2).font = Font(bold=True)

    # ── Aba Histórico ──
    ws4 = wb.create_sheet("Histórico")
    escrever_cabecalho(ws4, ["Mês", "Receita (R$)", "Despesa (R$)", "Saldo (R$)"])
    historico = [
        ("2025-11", 56800, 41200, 15600),
        ("2025-12", 58200, 44800, 13400),
        ("2026-01", 57100, 43600, 13500),
        ("2026-02", 55900, 40100, 15800),
        ("2026-03", 59300, 45200, 14100),
        ("2026-04", 57600, 42180, 15420),
    ]
    for r, row in enumerate(historico, 2):
        for c, val in enumerate(row, 1):
            ws4.cell(r, c, val)

    saida = DESTINO / "balancete_2026-05.xlsx"
    wb.save(saida)
    print(f"[OK] Arquivo de exemplo criado: {saida.relative_to(ROOT)}")


if __name__ == "__main__":
    criar_xlsx()
