"""
Adapter Empresa A — planilha com abas separadas:
  - Aba "Receitas":   col A=descrição, col B=previsto, col C=realizado
  - Aba "Despesas":   col A=categoria,  col B=valor
  - Aba "Resumo":     células fixas com saldo anterior e inadimplência
  - Aba "Histórico":  tabela com colunas Mês | Receita | Despesa | Saldo

Ajuste as referências de colunas/linhas conforme o arquivo real.
"""
from pathlib import Path
import openpyxl
from adapters.base import AdapterBase, DadosFinanceiros


class AdapterEmpresaA(AdapterBase):

    def ler_xlsx(self, caminho: Path, mes_referencia: str) -> DadosFinanceiros:
        wb = openpyxl.load_workbook(caminho, data_only=True)
        dados = DadosFinanceiros(
            condominio_id=self.config.get("id", ""),
            mes_referencia=mes_referencia,
        )

        # ---------- Aba Receitas ----------
        if "Receitas" in wb.sheetnames:
            ws = wb["Receitas"]
            prevista = 0.0
            realizada = 0.0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                prevista  += float(row[1] or 0)
                realizada += float(row[2] or 0)
            dados.receita_prevista  = prevista
            dados.receita_realizada = realizada

        # ---------- Aba Despesas ----------
        if "Despesas" in wb.sheetnames:
            ws = wb["Despesas"]
            total = 0.0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                categoria = str(row[0])
                valor = float(row[1] or 0)
                total += valor
                dados.categorias_despesa[categoria] = (
                    dados.categorias_despesa.get(categoria, 0) + valor
                )
            dados.despesa_total = total

        # ---------- Aba Resumo ----------
        if "Resumo" in wb.sheetnames:
            ws = wb["Resumo"]
            # Saldo anterior em B2, inadimplência valor em B4, % em B5
            dados.saldo_anterior          = self._valor_celula(ws, 2, 2)
            dados.inadimplencia_valor     = self._valor_celula(ws, 4, 2)
            dados.inadimplencia_percentual = self._valor_celula(ws, 5, 2)
            dados.unidades_inadimplentes  = int(self._valor_celula(ws, 6, 2))
            dados.total_unidades          = self.config.get("unidades", 0)

        # ---------- Aba Histórico ----------
        if "Histórico" in wb.sheetnames:
            ws = wb["Histórico"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                dados.historico_meses.append({
                    "mes":     str(row[0]),
                    "receita": float(row[1] or 0),
                    "despesa": float(row[2] or 0),
                    "saldo":   float(row[3] or 0),
                })

        dados.saldo_atual = self.calcular_saldo(dados)
        return dados
